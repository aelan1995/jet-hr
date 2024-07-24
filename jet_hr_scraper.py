from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.chrome.options import Options
import logging
import time
from datetime import datetime
import openpyxl
import re
from selenium.common.exceptions import JavascriptException
from openpyxl.styles import Font
import os
from openpyxl import Workbook
from bs4 import BeautifulSoup
from fake_useragent import UserAgent


# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

headers = [
            "Datetime", "Address URL", "Company Name", "Updated", "Business Establishment",
            "Location", "Fiscal Code", "Legal Form", "Internet Site Name", "NACE Code",
            "Sector", "Legal Representative", "Number of Employees Range", "EM Text",
            "LinkedIn Profile URL", "Presentation", "Competitors"
]

ua = UserAgent()
user_agent = ua.random

def get_text_after_label(label):
    try:
        label_element = soup.find('div', string=re.compile(label))
        return label_element.find_next_sibling('div').text.strip()
    except:
        return ""

chrome_options = Options()
# chrome_options.add_argument("--headless")  # Enable headless mode
# chrome_options.add_argument("--disable-gpu")
# chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument(f"user-agent={user_agent}")





# Set up Chrome options

# Execute the function to click all "Find out more" buttons
try:
    # Initialize the Chrome WebDriver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    wait = WebDriverWait(driver, 400)
    # Record the start time
    start_time = time.time()
    logging.info("Starting to load the web page.")

    # Open the desired URL directly
    driver.get("https://startup.registroimprese.it/isin/home#")

    driver.set_page_load_timeout(300)


    # Wait for the advanced search link to be clickable
    logging.info("Waiting for the 'advanced search' link to become clickable.")
    advanced_search_link = wait.until(
        EC.element_to_be_clickable((By.ID, "filtroAvanzatoLnk"))
    )

    # Click the advanced search link
    logging.info("'Advanced search' link is clickable. Clicking the link.")
    advanced_search_link.click()

    # Locate the parent div by its id
    # Click the region search link
    logging.info("Finding ID's")
    parent_div = driver.find_element(By.ID, "valueRegione")


    # Clear the inner HTML of the parent div
    children = parent_div.find_elements(By.XPATH, './*')
    get_id = []
    for child in children:
        # Find all grandchildren of each child
        grandchildren = child.find_elements(By.XPATH, './*')
        for grandchild in grandchildren:
            get_id.append(grandchild.get_attribute("id"))
            supergrandchildren = grandchild.find_elements(By.XPATH, './*')
            for supergrandchild in supergrandchildren:
                get_id.append(supergrandchild.get_attribute("id"))

    logging.info("Click Region")
    region_click = wait.until(
        EC.element_to_be_clickable((By.ID, f"{get_id[0]}"))
    )

    children1 = parent_div.find_elements(By.CSS_SELECTOR, 'div[data-value]')


    remove_first_data = 0
    for child1 in children1:
        data_value = child1.get_attribute('data-value')

        print(f'Clicking element with data-value: {data_value}')

        if remove_first_data == 0:
            try:
                try:
                    # Scroll the element into view
                    driver.execute_script("arguments[0].scrollIntoView(true);", child1)
                    time.sleep(0.1)  # Allow some time for the scrolling

                    # Click the element using JavaScript to avoid interactability issues
                    driver.execute_script("arguments[0].click();", child1)

                except ElementNotInteractableException:
                    print(f"Element with data-value {data_value} not interactable.")
            except ElementNotInteractableException:
                    print(f"Element with data-value {data_value} not interactable.")

        else:
            try:
                delete_icon = driver.find_element(By.XPATH, "//i[@class='delete icon']")
                print(f'Clicking element with data-value: {data_value}')
                try:
                    # Scroll the element into view
                    driver.execute_script("arguments[0].scrollIntoView(true);", delete_icon)
                    time.sleep(0.1)  # Allow some time for the scrolling

                    # Click the element using JavaScript to avoid interactability issues
                    driver.execute_script("arguments[0].click();", delete_icon)

                except ElementNotInteractableException:
                    print(f"Element with data-value {data_value} not interactable.")
            except ElementNotInteractableException:
                    print(f"Element with data-value {data_value} not interactable.")
            except Exception as e:
                print(f"An error occurred: {e}")

            try:
                try:
                    # Scroll the element into view
                    driver.execute_script("arguments[0].scrollIntoView(true);", child1)
                    time.sleep(0.1)  # Allow some time for the scrolling

                    # Click the element using JavaScript to avoid interactability issues
                    driver.execute_script("arguments[0].click();", child1)

                except ElementNotInteractableException:
                    print(f"Element with data-value {data_value} not interactable.")
            except ElementNotInteractableException:
                print(f"Element with data-value {data_value} not interactable.")

        # Wait for the checkboxes to be present
        try:
            checkboxes = wait.until(
                EC.presence_of_all_elements_located((By.XPATH, "//input[@type='checkbox']"))
            )
            # Use JavaScript to check each checkbox
            for checkbox in checkboxes[:2]:
                driver.execute_script("arguments[0].checked = true;", checkbox)
        except:
            print("No Check Boxes")


        #   # Example action: print the text of the span element
        # try:
        #     get_region = wait.until(
        #         EC.presence_of_element_located((By.XPATH, f"//a[@class='ui label transition visible' and @data-value='{data_value}']/span"))
        #     )
        #     # Print the text of the span element to verify
        #     get_region = get_region.text
        # except Exception as e:
        #     print(f"An error occurred: {e}")

        file_path = rf"D:\Documents\SideProjectFiles\Upwork\jet-hr\sample_output.xlsx"
        file_exists = os.path.exists(file_path)
        if not file_exists:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)
        else:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active


        # Wait for the element to be present
        search_button = wait.until(
            EC.presence_of_element_located((By.CLASS_NAME, "searchBtnVetrina"))
        )

        try:
            # Click the element using JavaScript
            driver.execute_script("arguments[0].click();", search_button)
        except JavascriptException as e:
            print(f"JavaScript click failed: {e}")



        wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, 'ui button') and contains(@class, 'rounded') and contains(@class, 'black') and @style='background-color: #525252;']")))


        # Find all div elements with the specified class and text
        buttons = driver.find_elements(By.XPATH, "//div[contains(@class, 'ui button') and contains(@class, 'rounded') and contains(@class, 'black') and @style='background-color: #525252;']")
        num_buttons = len(buttons)
        index = 1

        while index < num_buttons:
            # Find the buttons again after each navigation back
            try:
                buttons = driver.find_elements(By.XPATH, "//div[contains(@class, 'ui button') and contains(@class, 'rounded') and contains(@class, 'black') and @style='background-color: #525252;']")
                button_div = buttons[index]
                print(f"Clicking button {index}/{num_buttons}")
                # Scroll to the element
                driver.execute_script("arguments[0].scrollIntoView(true);", button_div)
                driver.execute_script("arguments[0].click();", button_div)
            except:
                # Wait for the page to load (optional, depending on your needs)
                driver.implicitly_wait(10)

                # Refresh the page
                driver.refresh()

                # Wait for a few seconds to observe the refresh (optional)
                driver.implicitly_wait(5)

                # Navigate back to the previous page
                driver.back()

                # Wait for a few seconds to observe the navigation (optional)
                driver.implicitly_wait(5)
                buttons = driver.find_elements(By.XPATH, "//div[contains(@class, 'ui button') and contains(@class, 'rounded') and contains(@class, 'black') and @style='background-color: #525252;']")
                button_div = buttons[index]
                # Scroll to the element
                driver.execute_script("arguments[0].scrollIntoView(true);", button_div)
                driver.execute_script("arguments[0].click();", button_div)

            time.sleep(10 * index)


            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')

            current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Using `string` instead of `text`
            try:
                company_name = soup.find('span', {'id': re.compile("companyNameForGA")}).string.strip()
            except:
                company_name = ""

            # Extract the header text
            header = soup.find('h2', class_='ui header').get_text(strip=True)

            # Extract the updated date
            updated_date_text = soup.find('div', class_='twelve wide column').get_text(strip=True)
            updated = updated_date_text.split()[-1]

            business_establishment = get_text_after_label("Business Establishment")
            location = get_text_after_label("Location")
            fiscal_code = get_text_after_label("Fiscal code")
            legal_form = get_text_after_label("Legal form")
            nace_code = get_text_after_label("NACE Code")
            sector = get_text_after_label("Sector")

            try:
                number_of_employees_range_element = soup.find('div', string=re.compile("Number of Employees Range"))
                number_of_employees_range = number_of_employees_range_element.find_next_sibling('div').find('span').string.strip() if number_of_employees_range_element else ""
            except:
                number_of_employees_range = ""

            try:
                email_url_elements = soup.find_all('a', href=re.compile('mailto:'))
                em_el = next((email_element for email_element in email_url_elements if email_element.string), None)
                em_url = em_el['href']
                em_text = em_el.string.strip()
            except:
                em_url = ""
                em_text = ""


            # Finding the span element with text matching "legal representative"
            try:
                span_element_get = soup.find('span', string=lambda x: x and 'legal representative' in x)

                # Extracting the text content
                text_content_get = span_element_get.get_text()

                # Splitting the text to get the name
                parts_get= text_content_get.split("legal representative ")[1].split(" on")
                legal_representative= parts_get[0].strip()

            except:
                legal_representative = ""
            try:
                presentation_an = soup.find('div', id='presentazioneAnchor')
                presentation = presentation_an.find('div', class_='twocolumntext').text.strip() if presentation_an else None
            except:
                presentation = ""

            try:
                competitors_details = []

                # Find the element containing the text "COMPETITORS"
                competitors_header = soup.find(string="COMPETITORS")

                if competitors_header:
                    # Get the parent div containing the header
                    parent_div = competitors_header.find_parent('div', class_='nobordered attached segment')
                    if parent_div:
                        # Extract text from the parent segment
                        competitors_text = parent_div.get_text(separator="\n", strip=True)
                        competitors_details.append(competitors_text)

                # Print the extracted competitors' details
                competitors = competitors_details[1]
            except:
                competitors = ""

            try:
                linkedin_profile_url = soup.find('a', href=re.compile('https://www.linkedin.com/company'))['href']
            except:
                linkedin_profile_url = ""

            try:
                internet_site_name_element = soup.find('div', string=re.compile('Internet site'))
                internet_site_name = internet_site_name_element.find_next_sibling('div').find('a').string.strip() if internet_site_name_element else ""
            except:
                internet_site_name = ""
            current_address_url = driver.current_url

            # Collect all data
            data = [
                current_datetime, current_address_url, company_name, updated, business_establishment,
                location, fiscal_code, legal_form, internet_site_name, nace_code, sector, legal_representative,
                number_of_employees_range, em_text, linkedin_profile_url, presentation, competitors, ""
            ]

            next_row = sheet.max_row + 1

            # Write the data to the next empty row
            for col_num, value in enumerate(data, 1):
                cell = sheet.cell(row=next_row, column=col_num, value=value)
                if value in [current_address_url, linkedin_profile_url, em_text, internet_site_name]:
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
                    cell.font = Font(color="0000FF", underline="single")

            workbook.save(file_path)

            print(f"URL after clicking button {index}: {current_address_url} at {current_datetime}")

            if index == 9:

                try:
                    driver.back()
                    time.sleep(5 * index)
                    next_page = wait.until(EC.presence_of_element_located((By.XPATH, "//a[@rel='next' and @style='padding: 0.5em 0.75em' and @title='Go to next page']")))
                    driver.execute_script("arguments[0].scrollIntoView(true);", next_page)
                    driver.execute_script("arguments[0].click();", next_page)
                    time.sleep(5 * index)
                    index = 1
                except:
                    remove_first_data = 1
                    index = 1
                    driver.back()
                    time.sleep(20)
                    break

            else:
                driver.back()
                time.sleep(5 * index)
                index += 1


        end_time = time.time()
        # Calculate the total time taken
        total_time = end_time - start_time
        # Print the total time taken
        logging.info(f"Time taken to access and load the web page: {total_time:.2f} seconds")



except Exception as e:
    logging.error(f"An error occurred: {e}", exc_info=True)
    end_time = time.time()
        # Calculate the total time taken
    total_time = end_time - start_time
        # Print the total time taken
    logging.info(f"Time taken to access and load the web page: {total_time:.2f} seconds")

finally:
    # Close the browser
    driver.quit()
    logging.info("Browser closed.")

