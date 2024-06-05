from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementNotInteractableException
import logging
import time
from datetime import datetime
import openpyxl
import traceback
import re

from openpyxl import Workbook
from openpyxl.styles import Font




# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# Function to get text after label
def get_text_after_label(label):
    print(f"Looking for label: {label}")  # Debugging line
    try:
        # Use a more general XPath
        element = wait.until(EC.presence_of_element_located((By.XPATH, f'//div[contains(., "{label}")]/following-sibling::div')))
        element_text = element.text
        print(f"Found text: {element_text}")  # Debugging line
        return element_text
    except Exception as e:
        print(f"Failed to find label {label}: {e}")
        print("Page HTML around the label:")
        try:
            surrounding_html = driver.find_element(By.XPATH, f'//div[contains(., "{label}")]').get_attribute('outerHTML')
            print(surrounding_html)
        except:
            print("Couldn't find the surrounding HTML for the label.")
        return None

# Initialize the Chrome WebDriver
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

# Record the start time
start_time = time.time()
logging.info("Starting to load the web page.")

# Open the desired URL directly
driver.get("file:///D:/Documents/SideProjectFiles/Upwork/jet-hr/jet-hr/sample3.html")

# Wait until elements are present
wait = WebDriverWait(driver, 20)

# Scrape the necessary data
current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
current_address_url = driver.current_url

# Company name
company_name = wait.until(EC.presence_of_element_located((By.XPATH, '//span[contains(@id, "companyNameForGA")]'))).text

# Updated date
updated_element = wait.until(EC.presence_of_element_located((By.XPATH, '//h2[contains(text(), "BUSINESS REGISTER INFORMATION")]')))
updated_text = driver.execute_script("return arguments[0].nextSibling.nodeValue;", updated_element).strip()
updated = updated_text.split("Updated")[1].strip()


business_establishment = get_text_after_label("Business Establishment")
location = get_text_after_label("Location")
fiscal_code = get_text_after_label("Fiscal code")
legal_form = get_text_after_label("Legal form")
internet_site = wait.until(EC.presence_of_element_located((By.XPATH, '//div[contains(text(), "Internet site")]/following-sibling::div/a'))).text
nace_code = get_text_after_label("NACE Code")
sector = get_text_after_label("Sector")

number_of_employees_range = wait.until(EC.presence_of_element_located((By.XPATH, '//div[contains(text(), "Number of Employees Range")]/following-sibling::div/span'))).text
linkedin_profile_url = wait.until(EC.presence_of_element_located((By.XPATH, '//a[contains(@href, "https://www.linkedin.com/company")]'))).get_attribute('href')

wait = WebDriverWait(driver, 10)
email_url_element = driver.find_element(By.XPATH, "//div//a[contains(@href, 'mailto:')]")

# Find all elements containing the email URLs
email_url_elements = driver.find_elements(By.XPATH, "//a[contains(@href, 'mailto:')]")

# Extract and print the text content of each element
for email_element in email_url_elements:
    email_text = email_element.text
    print("Email Text:", email_text)
email_href = email_url_element.get_attribute('href')

[email_element.text for index, email_element in enumerate(email_url_elements) if email_element.text]

# Print the href attribute of the <a> tag inside the next sibling <div>





try:

    legal_representative_element = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'legal representative')]")))
    name = legal_representative_element.text.split(":")[-1].strip()  # Extracting the name after the colon and trimming spaces
    pattern = r"by the legal representative (.*?) on"
    match = re.search(pattern, name)
    legal_representative = match.group(1)

    h4_element = driver.find_element(By.XPATH, "//h4[contains(@class, 'header')]//span[text()='PRESENTATION']")
    parent_h4 = h4_element.find_element(By.XPATH, "./ancestor::h4")
    presentation= parent_h4.find_element(By.XPATH, "following-sibling::div[2]").text

    h4_element_2 = driver.find_element(By.XPATH, "//h4[contains(@class, 'header')]//span[text()='COMPETITORS']")
    parent_h4_2 = h4_element_2.find_element(By.XPATH, "./ancestor::h4")
    competitors= parent_h4_2.find_element(By.XPATH, "following-sibling::div[2]").text

except:
    # Assign a blank string if the section is not found
    presentation = ""
    competitors = ""
    legal_representative = ""

# Close the WebDriver
driver.quit()

import json
# Data dictionary
data = {
    "current_datetime": current_datetime,
    "current_address_url": current_address_url,
    "company_name": company_name,
    "updated": updated,
    "business_establishment": business_establishment,
    "location": location,
    "fiscal_code": fiscal_code,
    "legal_form": legal_form,
    "internet_site": internet_site,
    "nace_code": nace_code,
    "sector": sector,
    "legal_representative": legal_representative,
    "number_of_employees_range": number_of_employees_range,
    "email": email_text,
    "linkedin_profile_url": linkedin_profile_url,
    "presentation": presentation,
    "competitors": competitors
}

# Print as pretty JSON
print(json.dumps(data, indent=4))


# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Define the URL and the display text
url = linkedin_profile_url
display_text = linkedin_profile_url

# Add the hyperlink to a cell
cell = ws.cell(row=1, column=1, value=display_text)
cell.hyperlink = url
cell.style = "Hyperlink"
cell.font = Font(color="0000FF", underline="single")

cell = ws.cell(row=2, column=2, value=email_text)
cell.hyperlink = email_href
cell.style = "Hyperlink"
cell.font = Font(color="0000FF", underline="single")

# Save the workbook
wb.save("example.xlsx")









